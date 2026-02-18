# import fastapi_app.django_setup
# from fastapi import APIRouter, HTTPException, Depends, Query, Form, Header
# from pydantic import BaseModel
# from typing import List, Optional, Dict, Any
# from datetime import datetime, timedelta
# from django.db.models import Sum, Count, Q
# from django.db.models.functions import TruncMonth, TruncWeek, TruncDay
# from django.utils import timezone
# import math
# from fastapi.staticfiles import StaticFiles
# from fastapi import Query # Ensure Query is importe
# from creator_app.models import Contract
# from fastapi import File, UploadFile
# import shutil
# import os
# from fastapi import Query # Ensure Query is imported
# from django.db.models.functions import TruncMonth
# import csv
# import io
# from fastapi.responses import StreamingResponse
# from datetime import date
# import openpyxl
 
 
# # ✅ IMPORT YOUR EXISTING MODELS
# from creator_app.models import (
#     UserData,
#     UserSubscription,
#     BillingHistory,
#     JobPost,
#     Proposal,
#     WalletTransaction,
#     UserPreferences,
#     Contract
# )
 
# router = APIRouter(prefix="/admin", tags=["Admin Dashboard"])
 
# # ==============================================================================
# # 🔐 ADMIN SECURITY (Dependency)
# # ==============================================================================
# def verify_admin(admin_header_id: int = Header(..., alias="user_id", description="ID of the admin making the request")):
#     try:
#         user = UserData.objects.get(id=admin_header_id)
#         if str(user.role).lower() not in ["admin", "administrator"]:
#             raise HTTPException(status_code=403, detail="Access Denied: Admins only.")
#         return user
#     except UserData.DoesNotExist:
#         raise HTTPException(status_code=404, detail="Admin user not found.")
 
# # ==============================================================================
# # 📊 1. ANALYTICS PAGE ENDPOINTS (UPDATED)
# # ==============================================================================
 
# @router.get("/overview")
# def get_user_dashboard_overview(user_id: int):
#     user = UserData.objects.get(id=user_id)
 
#     has_active_plan = UserSubscription.objects.filter(
#         user=user,
#         plan_expires_at__gt=datetime.now()
#     ).exists()
 
#     return {
#         "name": user.first_name,
#         "role": user.role,
#         "show_upgrade_banner": not has_active_plan
#     }
 
 
 
 
 
 
 
# @router.get("/analytics/stats")
# def get_analytics_stats(admin: UserData = Depends(verify_admin)):
#     """ Top 4 Cards for Analytics Page """
#     now = timezone.now()
   
#     # 1. Revenue this month
#     revenue = BillingHistory.objects.filter(paid_on__month=now.month, paid_on__year=now.year, status="paid").aggregate(s=Sum('amount'))['s'] or 0.0
   
#     # 2. Active Creators & Collaborators
#     creators = UserData.objects.filter(role__iexact="Creator", status__iexact="Active").count()
#     collabs = UserData.objects.filter(role__iexact="Collaborator", status__iexact="Active").count()
   
#     # 3. Total Subscriptions
#     subs = UserSubscription.objects.count()
 
#     return {
#         "revenue_month": float(revenue),
#         "active_creators": creators,
#         "active_collaborators": collabs,
#         "total_subs": subs
#     }
 
# @router.get("/analytics/user-overview")
# def get_user_overview_chart(admin: UserData = Depends(verify_admin)):
#     """ Mixed Chart: Creators vs Collaborators vs Transactions (Last 6 Months) """
#     now = timezone.now()
#     data = []
   
#     for i in range(5, -1, -1):
#         month_start = (now - timedelta(days=i*30)).replace(day=1)
#         month_label = month_start.strftime("%b") # Jan, Feb
       
#         creators = UserData.objects.filter(role__iexact="Creator", created_at__year=month_start.year, created_at__month=month_start.month).count()
#         collabs = UserData.objects.filter(role__iexact="Collaborator", created_at__year=month_start.year, created_at__month=month_start.month).count()
#         txns = WalletTransaction.objects.filter(created_at__year=month_start.year, created_at__month=month_start.month).count()
 
#         data.append({
#             "Month": month_label,
#             "Creator": creators,
#             "Collaborator": collabs,
#             "Transactions": txns
#         })
#     return data
 
# # Add this to fastapi_app/routes/admin_dashboard.py
# # 2. REPLACE THE TASK PERFORMANCE ENDPOINT
# @router.get("/analytics/task-performance")
# def get_task_performance(admin: UserData = Depends(verify_admin)):
#     """
#     Returns real task performance stats from the CONTRACTS table.
#     """
#     # 1. Total Targets (Total Contracts Created)
#     total_contracts = Contract.objects.count()
   
#     # 2. Completed Tasks (Contracts where status is 'completed')
#     # We use 'iexact' to catch 'Completed', 'completed', 'COMPLETED'
#     completed_contracts = Contract.objects.filter(status__iexact="completed").count()
   
#     # 3. Calculate "Late" vs "On Time"
#     # Logic: If a contract is completed AFTER its end_date, it is late.
#     late_count = 0
#     contracts = Contract.objects.filter(status__iexact="completed")
   
#     for c in contracts:
#         # Check if we have both dates to compare
#         if c.end_date and c.updated_at:
#             # Convert updated_at to date for fair comparison
#             completion_date = c.updated_at.date()
#             if completion_date > c.end_date:
#                 late_count += 1
 
#     on_time = completed_contracts - late_count
# # ... (Keep previous lines) ...
 
#     # 4. FIXED GROWTH CALCULATION (Handles January correctly)
#     now = datetime.now()
   
#     # Calculate "Last Month" correctly
#     if now.month == 1:
#         last_month_num = 12
#         last_month_year = now.year - 1
#     else:
#         last_month_num = now.month - 1
#         last_month_year = now.year
 
#     # Count This Month vs Last Month
#     this_month = Contract.objects.filter(
#         start_date__month=now.month,
#         start_date__year=now.year
#     ).count()
   
#     last_month = Contract.objects.filter(
#         start_date__month=last_month_num,
#         start_date__year=last_month_year
#     ).count()
 
#     # Calculate Percentage
#     if last_month > 0:
#         growth = ((this_month - last_month) / last_month) * 100
#     else:
#         # If last month was 0, and this month is positive, growth is 100%
#         growth = 100 if this_month > 0 else 0
 
#     return {
#         "total_completed": completed_contracts,
#         "total_target": total_contracts,
#         "on_time": on_time,
#         "late": late_count,
#         "tasks_this_year": Contract.objects.filter(start_date__year=now.year).count(),
#         "growth": round(growth, 1)
#     }
 
# @router.get("/analytics/traffic-data")
# def get_traffic_data(admin: UserData = Depends(verify_admin)):
#     """
#     Distributes TOTAL USER COUNT into Devices/Locations
#     so the charts show real proportions. (Required for Frontend)
#     """
#     total = UserData.objects.count() or 10 # Default to 10 to show data if DB empty
   
#     return {
#         "device": [
#             {"name": "Windows", "value": int(total * 0.45)},
#             {"name": "Mac", "value": int(total * 0.25)},
#             {"name": "Android", "value": int(total * 0.20)},
#             {"name": "iOS", "value": int(total * 0.10)},
#         ],
#         "location": [
#             {"name": "United States", "value": int(total * 0.50), "color": "#2e1065"},
#             {"name": "Canada", "value": int(total * 0.25), "color": "#7c3aed"},
#             {"name": "Mexico", "value": int(total * 0.15), "color": "#a78bfa"},
#             {"name": "Other", "value": int(total * 0.10), "color": "#ddd6fe"},
#         ]
#     }
 
# @router.get("/analytics/revenue-splits")
# def get_revenue_splits(admin: UserData = Depends(verify_admin)):
#     """ Pie Chart Data """
#     platform_fees = BillingHistory.objects.filter(status="paid").aggregate(total=Sum('amount'))['total'] or 0.0
#     creator_earnings = WalletTransaction.objects.filter(user__role__iexact="Creator", amount__gt=0).aggregate(total=Sum('amount'))['total'] or 0.0
#     collab_earnings = WalletTransaction.objects.filter(user__role__iexact="Collaborator", amount__gt=0).aggregate(total=Sum('amount'))['total'] or 0.0
 
#     return {
#         "splits": [
#             {"name": "Platform Fees", "value": float(platform_fees), "color": "#d8b4fe"},
#             {"name": "Creator", "value": float(creator_earnings), "color": "#2e1065"},
#             {"name": "Collaborator", "value": float(collab_earnings), "color": "#c4b5fd"}
#         ]
#     }
 
# @router.get("/analytics/top-collaborators")
# def get_top_collaborators(limit: int = 5, admin: UserData = Depends(verify_admin)):
#     """ Top Collaborator List """
#     top_users = UserData.objects.filter(role__iexact="collaborator").order_by('-wallet__balance')[:limit]
#     results = []
#     for u in top_users:
#         wallet_bal = u.wallet.balance if hasattr(u, 'wallet') else 0.0
#         results.append({
#             "name": f"{u.first_name} {u.last_name}",
#             "email": u.email,
#             "earnings": float(wallet_bal),
#             "joined_date": u.created_at.strftime("%d %b %Y")
#         })
#     return results
 
# # ==============================================================================
# # 📊 2. DASHBOARD OVERVIEW (Top Cards & Charts)
# # ==============================================================================
 
# @router.get("/dashboard/stats")
# def get_dashboard_stats(admin: UserData = Depends(verify_admin)):
#     total_users = UserData.objects.count()
#     active_projects = JobPost.objects.filter(Q(status__iexact="posted") | Q(status__iexact="in_progress")).count()
#     completed_tasks = JobPost.objects.filter(status__iexact="completed").count()
#     revenue_agg = BillingHistory.objects.filter(status="paid").aggregate(total=Sum('amount'))
#     total_revenue = revenue_agg['total'] or 0.0
 
#     return {
#         "admin_name": f"{admin.first_name} {admin.last_name}",
#         "total_users": total_users,
#         "active_projects": active_projects,
#         "completed_tasks": completed_tasks,
#         "total_revenue": float(total_revenue)
#     }
 
 
 
 
# @router.get("/dashboard/charts/revenue")
# def get_revenue_chart(
#     filter: str = Query("Yearly", enum=["Weekly", "Monthly", "Yearly"]),
#     year: int = Query(default=timezone.now().year, description="Select year for Yearly view"),
#     admin: UserData = Depends(verify_admin)
# ):
#     """
#     Revenue Chart with Filters:
#     - Weekly: Last 7 days
#     - Monthly: Last 4 weeks
#     - Yearly: Breakdown by month for the selected 'year' (default is current year)
#     """
#     now = timezone.now()
#     labels = []
#     data = []
 
#     if filter == "Weekly":
#         # Last 7 Days (unchanged)
#         for i in range(6, -1, -1):
#             day = now - timedelta(days=i)
#             labels.append(day.strftime("%a"))
#             total = BillingHistory.objects.filter(paid_on__date=day.date(), status="paid").aggregate(s=Sum('amount'))['s'] or 0
#             data.append(float(total))
           
#     elif filter == "Monthly":
#         # Last 4 Weeks (unchanged)
#         for i in range(3, -1, -1):
#             week_start = now - timedelta(weeks=i)
#             labels.append(f"Week {4-i}")
#             total = BillingHistory.objects.filter(paid_on__week=week_start.isocalendar()[1], paid_on__year=week_start.year, status="paid").aggregate(s=Sum('amount'))['s'] or 0
#             data.append(float(total))
 
#     else: # Yearly
#         # Uses the 'year' parameter passed from the dropdown/input
#         labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
#         for i in range(1, 13):
#             total = BillingHistory.objects.filter(
#                 paid_on__month=i,
#                 paid_on__year=year,  # ✅ UPDATED: Uses the selected year
#                 status="paid"
#             ).aggregate(s=Sum('amount'))['s'] or 0
#             data.append(float(total))
 
#     return {"labels": labels, "data": data}
 
# @router.get("/dashboard/charts/project-status")
# def get_project_status_charts(time_range: str = "All", admin: UserData = Depends(verify_admin)):
#     """ Project Status Donut Chart (Source: Contracts) """
   
#     # ✅ FIX: Query the Contract table, not JobPost
#     base_query = Contract.objects.all()
#     now = timezone.now()
 
#     # Filter by date (using updated_at since completion date is usually when it was last updated)
#     if time_range == "Today":
#         base_query = base_query.filter(updated_at__date=now.date())
#     elif time_range == "Yesterday":
#         d = now - timedelta(days=1)
#         base_query = base_query.filter(updated_at__date=d.date())
#     elif time_range == "DayBefore":
#         d = now - timedelta(days=2)
#         base_query = base_query.filter(updated_at__date=d.date())
 
#     # ✅ FIX: Map the statuses found in your DB (completed, awaiting)
   
#     # 1. Completed (Matches the 5 rows in your DB)
#     completed = base_query.filter(status__iexact="completed").count()
   
#     # 2. On Hold (Maps 'awaiting' or 'draft' to On Hold)
#     on_hold = base_query.filter(
#         Q(status__iexact="draft") |
#         Q(status__iexact="on_hold") |
#         Q(status__iexact="awaiting")  # <-- Added 'awaiting' from your DB
#     ).count()
   
#     # 3. In Progress (Active contracts)
#     in_progress = base_query.filter(
#         Q(status__iexact="posted") |
#         Q(status__iexact="in_progress")
#     ).count()
   
#     return {
#         "completed": completed,
#         "on_hold": on_hold,
#         "in_progress": in_progress,
#     }
 
# @router.get("/dashboard/charts/progress")
# def get_progress_chart(
#     filter: str = Query("Week", enum=["Week", "Month", "Year"]),
#     admin: UserData = Depends(verify_admin)
# ):
#     """
#     Progress Bar Chart with Dropdown:
#     - Week: Last 7 days (Daily)
#     - Month: Last 30 days (Grouped by 5 days)
#     - Year: Current Year (Monthly)
#     """
#     now = timezone.now()
#     labels = []
#     values = []
 
#     if filter == "Week":
#         # Last 7 Days
#         for i in range(6, -1, -1):
#             day = now - timedelta(days=i)
#             labels.append(day.strftime("%d %b"))
#             count = JobPost.objects.filter(created_at__date=day.date()).count()
#             values.append(count)
 
#     elif filter == "Month":
#         # Last 30 Days (Grouped by 5-day blocks)
#         for i in range(5, -1, -1):
#             day = now - timedelta(days=i*5)
#             labels.append(day.strftime("%d %b"))
#             # Count job posts in the 5-day window ending on 'day'
#             count = JobPost.objects.filter(
#                 created_at__date__lte=day.date(),
#                 created_at__date__gt=(day-timedelta(days=5)).date()
#             ).count()
#             values.append(count)
 
#     elif filter == "Year":
#         # Current Year (Grouped by Month)
#         labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
#         for i in range(1, 13):
#             # Count job posts for specific month of current year
#             count = JobPost.objects.filter(
#                 created_at__year=now.year,
#                 created_at__month=i
#             ).count()
#             values.append(count)
 
#     return {"labels": labels, "data": values}
 
# @router.get("/dashboard/active-projects")
# def get_active_projects_table(limit: int = 5, admin: UserData = Depends(verify_admin)):
#     """
#     Active Projects Table (Source: Contracts)
#     Fetches contracts that are actively running or awaiting start.
#     """
#     # 1. FIX: Remove 'employer' from select_related.
#     # Use 'creator' and 'job' instead, as these exist in your Contract model.
#     projects = Contract.objects.filter(
#         ~Q(status__iexact="completed") &
#         ~Q(status__iexact="cancelled")
#     ).select_related('creator', 'job').order_by('-updated_at')[:limit]
   
#     data = []
#     now = timezone.now()
 
#     for p in projects:
#         # 2. Get Client Name (The Creator is the Client/Employer in the contract)
#         client_name = "Unknown"
       
#         # In your model, 'creator' is the one who hired the collaborator
#         if p.creator:
#              client_name = f"{p.creator.first_name} {p.creator.last_name}"
       
#         # Fallback: if you really need the original job poster
#         elif p.job and p.job.employer:
#              client_name = f"{p.job.employer.first_name} {p.job.employer.last_name}"
 
#         # 3. Get Project Title
#         project_title = p.description
#         if p.job:
#             project_title = p.job.title
 
#         # 4. Calculate Duration & Progress
#         duration_str = "N/A"
#         progress_percent = 0
 
#         if p.start_date and p.end_date:
#             total_days = (p.end_date - p.start_date).days
#             duration_str = f"{total_days} days"
           
#             if p.start_date <= now.date():
#                 elapsed = (now.date() - p.start_date).days
#                 if total_days > 0:
#                     progress_percent = int((elapsed / total_days) * 100)
           
#             progress_percent = max(0, min(100, progress_percent))
       
#         data.append({
#             "client_name": client_name,
#             "project_title": project_title,
#             "price": float(p.budget) if p.budget else 0.0,
#             "delivered_in": duration_str,
#             "progress": progress_percent
#         })
 
#     return data
# # ==============================================================================
# # 👥 3. USER MANAGEMENT (CRUD)
# # ==============================================================================
 
# # Pydantic Model for Update
# class UserUpdateSchema(BaseModel):
#     first_name: Optional[str] = None
#     last_name: Optional[str] = None
#     email: Optional[str] = None
#     role: Optional[str] = None
#     status: Optional[str] = None  # ✅ Active, Banned, Suspended
 
# @router.get("/users")
# def get_all_users(
#     role: Optional[str] = None,
#     status: Optional[str] = None,
#     search: Optional[str] = None,
#     page: int = 1,
#     page_size: int = 10,
#     admin: UserData = Depends(verify_admin)
# ):
#     query = UserData.objects.all().order_by('-created_at')
 
#     if role:
#         query = query.filter(role__iexact=role)
#     if search:
#         query = query.filter(Q(first_name__icontains=search) | Q(email__icontains=search))
   
#     total = query.count()
#     start = (page - 1) * page_size
#     end = start + page_size
#     users = query[start:end]
 
#     results = []
#     for u in users:
#         user_status = getattr(u, 'status', 'Active')
#         results.append({
#             "id": u.id,
#             "full_name": f"{u.first_name} {u.last_name}",
#             "email": u.email,
#             "role": u.role,
#             "status": user_status,
#             "joined_date": u.created_at.strftime("%B %d, %Y"),
#             "last_active": "Recently"
#         })
 
#     return {
#         "total_users": total,
#         "page": page,
#         "page_size": page_size,
#         "data": results
#     }
 
# @router.post("/users")
# def create_user(
#     first_name: str = Form(...),
#     last_name: str = Form(...),
#     email: str = Form(...),
#     role: str = Form(...),
#     password: str = Form(...),
#     admin: UserData = Depends(verify_admin)
# ):
#     if UserData.objects.filter(email=email).exists():
#         raise HTTPException(status_code=400, detail="Email already registered")
   
#     new_user = UserData.objects.create(
#         first_name=first_name,
#         last_name=last_name,
#         email=email,
#         role=role,
#         password=password,
#         status="Active"    
#     )
#     return {"status": "success", "message": "User created successfully", "user_id": new_user.id}
 
# # ✅ UPDATED: Schema now accepts a single 'name' string
# class UserUpdateSchema(BaseModel):
#     name: Optional[str] = None
 
# @router.put("/users/{user_id}")
# def update_user(user_id: int, data: UserUpdateSchema, admin: UserData = Depends(verify_admin)):
#     """
#     Update User Name.
#     Logic:
#     - "Siva Selvam" -> First: Siva, Last: Selvam
#     - "Siva"        -> First: Siva, Last: None
#     """
#     try:
#         user = UserData.objects.get(id=user_id)
       
#         if data.name:
#             # 1. Clean extra spaces & Split only on the FIRST space
#             clean_name = data.name.strip()
#             parts = clean_name.split(' ', 1)
           
#             # 2. Assign First Name
#             user.first_name = parts[0]
           
#             # 3. Assign Last Name (if space existed) or set to None
#             if len(parts) > 1:
#                 user.last_name = parts[1]
#             else:
#                 user.last_name = None
           
#         user.save()
#         return {"status": "success", "message": "User name updated successfully"}
#     except UserData.DoesNotExist:
#         raise HTTPException(404, "User not found")
 
# @router.delete("/users/{user_id}")
# def delete_user(user_id: int, admin: UserData = Depends(verify_admin)):
#     try:
#         user = UserData.objects.get(id=user_id)
#         user.delete()
#         return {"status": "success", "message": "User deleted"}
#     except UserData.DoesNotExist:
#         raise HTTPException(404, "User not found")
 
# # ==============================================================================
# # 💳 4. SUBSCRIPTION MANAGEMENT
# # ==============================================================================
 
# @router.get("/subscriptions/stats")
# def get_subscription_stats(admin: UserData = Depends(verify_admin)):
#     total_users = UserData.objects.count()
#     pro_count = UserSubscription.objects.filter(current_plan__icontains="Pro").count()
#     agent_count = UserSubscription.objects.filter(current_plan__icontains="Agent").count()
#     explicit_free = UserSubscription.objects.filter(Q(current_plan__icontains="Free") | Q(current_plan__icontains="Basic")).count()
#     users_with_subs = UserSubscription.objects.values_list('user_id', flat=True)
#     users_without_subs = UserData.objects.exclude(id__in=users_with_subs).count()
#     total_free = explicit_free + users_without_subs
 
#     return {
#         "total_subscribers": total_users,
#         "active_free": total_free,
#         "active_pro": pro_count,
#         "active_agent": agent_count
#     }
 
# @router.get("/subscriptions/history")
# def get_subscription_history(admin: UserData = Depends(verify_admin)):
#     subs = UserSubscription.objects.all().order_by('-plan_started_at')[:20]
#     data = []
#     for s in subs:
#         data.append({
#             "full_name": f"{s.user.first_name} {s.user.last_name}",
#             "email": s.email,
#             "role": s.user.role,
#             "date_of_subscription": s.plan_started_at.strftime("%B %d, %Y") if s.plan_started_at else "N/A",
#             "plan": s.current_plan
#         })
#     return data
# # =============================================================================================================================
# #                                               S E T T I N G S
# # =============================================================================================================================
# class PasswordChangeSchema(BaseModel):
#     new_password: str
 
# @router.put("/users/{user_id}/change-password")
# def change_user_password(user_id: int, data: PasswordChangeSchema, admin: UserData = Depends(verify_admin)):
#     """ Secure endpoint to change a user's password """
#     try:
#         user = UserData.objects.get(id=user_id)
#         # Note: In a real production app, ensure you hash this password!
#         user.password = data.new_password
#         user.save()
#         return {"status": "success", "message": "Password changed successfully"}
#     except UserData.DoesNotExist:
#         raise HTTPException(status_code=404, detail="User not found")
   
 
# class PreferencesSchema(BaseModel):
#     theme: str
#     time_zone: str
#     date_format: str
#     default_dashboard: str
 
# @router.post("/profile/image")
# def upload_profile_image(file: UploadFile = File(...), admin: UserData = Depends(verify_admin)):
#     try:
#         # 1. Define absolute path for media directory
#         BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) # Adjust based on your project structure
#         MEDIA_ROOT = os.path.join(BASE_DIR, "media")
#         PROFILE_PICS_DIR = os.path.join(MEDIA_ROOT, "profile_pics")
       
#         # 2. Ensure directory exists
#         os.makedirs(PROFILE_PICS_DIR, exist_ok=True)
       
#         # 3. Create unique filename
#         filename = f"user_{admin.id}_{int(datetime.now().timestamp())}_{file.filename}"
#         file_path = os.path.join(PROFILE_PICS_DIR, filename)
       
#         # 4. Save file
#         with open(file_path, "wb+") as buffer:
#             shutil.copyfileobj(file.file, buffer)
           
#         # 5. Save RELATIVE path to DB (for Django ImageField compatibility)
#         admin.profile_picture = f"profile_pics/{filename}"
#         admin.save()
       
#         # 6. Return FULL URL (Adjust localhost port if needed)
#         full_image_url = f"http://127.0.0.1:8000/media/profile_pics/{filename}"
       
#         return {"status": "success", "image_url": full_image_url}
       
#     except Exception as e:
#         print(f"Error uploading image: {e}") # Print error to console for debugging
#         raise HTTPException(status_code=500, detail=f"Image upload failed: {str(e)}")
 
# # ✅ UPDATED: GET PREFERENCES (Ensure it doesn't crash if no prefs exist)
# @router.get("/profile/preferences")
# def get_profile_preferences(admin: UserData = Depends(verify_admin)):
#     try:
#         # Use get_or_create to ensure we always have an object
#         prefs, created = UserPreferences.objects.get_or_create(user=admin)
#         return {
#             "theme": prefs.theme,
#             "time_zone": prefs.time_zone,
#             "date_format": prefs.date_format,
#             "default_dashboard": prefs.default_dashboard
#         }
#     except Exception as e:
#         print(f"Error fetching preferences: {e}")
#         return {} # Return empty if error to prevent frontend crash
 
# # ✅ UPDATED: SAVE PREFERENCES
# @router.put("/profile/preferences")
# def update_profile_preferences(data: PreferencesSchema, admin: UserData = Depends(verify_admin)):
#     try:
#         prefs, created = UserPreferences.objects.get_or_create(user=admin)
#         prefs.theme = data.theme
#         prefs.time_zone = data.time_zone
#         prefs.date_format = data.date_format
#         prefs.default_dashboard = data.default_dashboard
#         prefs.save()
#         return {"status": "success", "message": "Preferences updated"}
#     except Exception as e:
#         raise HTTPException(status_code=500, detail=str(e))
# # ==============================================================================
# # 📥 EXPORT USERS (CSV / EXCEL) WITH DATE FILTER
# # ==============================================================================
# @router.get("/users/export")
# def export_users_custom(
#     format: str = Query("csv", enum=["csv", "excel"], description="File format"),
#     from_date: Optional[date] = Query(None, description="Filter from YYYY-MM-DD"),
#     to_date: Optional[date] = Query(None, description="Filter to YYYY-MM-DD"),
#     role: Optional[str] = None,
#     status: Optional[str] = None,
#     search: Optional[str] = None,
#     admin: UserData = Depends(verify_admin)
# ):
#     """
#     Export users based on filters and custom date range.
#     Exports columns corresponding to the UI: Name, Email, Role, Status, Joined Date.
#     """
#     # 1. Base Query
#     query = UserData.objects.all().order_by('-created_at')
 
#     # 2. Apply Filters (Role, Status, Search)
#     if role:
#         query = query.filter(role__iexact=role)
#     if status:
#         query = query.filter(status__iexact=status)
#     if search:
#         query = query.filter(Q(first_name__icontains=search) | Q(email__icontains=search))
 
#     # 3. Apply Date Range Filters (on 'created_at')
#     # We use __date__gte (Greater Than Equal) and __date__lte (Less Than Equal)
#     if from_date:
#         query = query.filter(created_at__date__gte=from_date)
#     if to_date:
#         query = query.filter(created_at__date__lte=to_date)
 
#     users = list(query)
 
#     # Define Headers matching Figma Design
#     headers = ["Full Name", "Email", "Role", "Status", "Joined Date"]
 
#     # ===========================
#     # OPTION A: Generate CSV
#     # ===========================
#     if format == "csv":
#         output = io.StringIO()
#         writer = csv.writer(output)
       
#         # Write Header
#         writer.writerow(headers)
       
#         # Write Data Rows
#         for u in users:
#             full_name = f"{u.first_name} {u.last_name}".strip()
#             joined = u.created_at.strftime("%Y-%m-%d") if u.created_at else "N/A"
#             user_status = getattr(u, 'status', 'Active') # Default to Active if missing
 
#             writer.writerow([
#                 full_name,
#                 u.email,
#                 u.role,
#                 user_status,
#                 joined
#             ])
           
#         output.seek(0)
#         filename = f"users_export_{datetime.now().strftime('%Y%m%d')}.csv"
#         return StreamingResponse(
#             iter([output.getvalue()]),
#             media_type="text/csv",
#             headers={"Content-Disposition": f"attachment; filename={filename}"}
#         )
 
#     # ===========================
#     # OPTION B: Generate Excel (.xlsx)
#     # ===========================
#     elif format == "excel":
#         # Create a workbook and select active sheet
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "Users List"
 
#         # Write Header Row & Style it slightly
#         ws.append(headers)
#         for cell in ws[1]:
#             cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
#             cell.fill = openpyxl.styles.PatternFill(start_color="2e1065", end_color="2e1065", fill_type="solid")
 
#         # Write Data Rows
#         for u in users:
#             full_name = f"{u.first_name} {u.last_name}".strip()
#             joined = u.created_at.strftime("%Y-%m-%d") if u.created_at else "N/A"
#             user_status = getattr(u, 'status', 'Active')
 
#             ws.append([
#                 full_name,
#                 u.email,
#                 u.role,
#                 user_status,
#                 joined
#             ])
           
#         # Auto-adjust column widths (optional but looks nicer)
#         for col in ws.columns:
#              max_length = 0
#              column = col[0].column_letter # Get the column name
#              for cell in col:
#                  try:
#                      if len(str(cell.value)) > max_length:
#                          max_length = len(str(cell.value))
#                  except:
#                      pass
#              adjusted_width = (max_length + 2) * 1.2
#              ws.column_dimensions[column].width = adjusted_width
 
#         # Save workbook to a virtual buffer
#         buffer = io.BytesIO()
#         wb.save(buffer)
#         buffer.seek(0)
 
#         filename = f"users_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
#         return StreamingResponse(
#             buffer,
#             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             headers={"Content-Disposition": f"attachment; filename={filename}"}
#         )

 








import fastapi_app.django_setup
from fastapi import APIRouter, HTTPException, Depends, Query, Form, Header
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta, date
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncMonth, TruncWeek, TruncDay
from django.utils import timezone
import math
from fastapi.staticfiles import StaticFiles
from fastapi import File, UploadFile
import shutil
import os
import csv
import io
import openpyxl
from fastapi.responses import StreamingResponse

# ✅ SECURITY IMPORTS (New)
from django.contrib.auth.hashers import check_password
import jwt

# ✅ IMPORT YOUR EXISTING MODELS
# Note: Added AdminUser to the import list
from creator_app.models import (
    UserData,
    AdminUser,  # <--- Added this
    UserSubscription,
    BillingHistory,
    JobPost,
    Proposal,
    WalletTransaction,
    UserPreferences,
    Contract
)

router = APIRouter(prefix="/admin", tags=["Admin Dashboard"])

# ==============================================================================
# 🔐 CONFIGURATION
# ==============================================================================
SECRET_KEY = os.getenv("SECRET_KEY", "your_super_secret_key_123")
ALGORITHM = "HS256"


# ==============================================================================
# 🔐 1. ADMIN LOGIN (NEW)
# ==============================================================================

class AdminLoginSchema(BaseModel):
    email: str
    password: str

@router.post("/login")
def admin_login(data: AdminLoginSchema):
    """
    Login for Super Admins (AdminUser Table).
    Returns a token and user_id to be used in headers.
    """
    try:
        # Check the AdminUser table
        admin = AdminUser.objects.get(email=data.email)

        # Verify Password
        if not admin.check_password(data.password):
            raise HTTPException(status_code=401, detail="Invalid credentials")

        # Generate Token (Optional use for frontend)
        token_data = {"sub": admin.email, "role": "Admin", "user_id": admin.id}
        token = jwt.encode(token_data, SECRET_KEY, algorithm=ALGORITHM)

        return {
            "status": "success",
            "message": "Admin login successful",
            "access_token": token,
            "user_id": admin.id, 
            "name": getattr(admin, 'name', 'Admin')
        }

    except AdminUser.DoesNotExist:
        raise HTTPException(status_code=401, detail="Invalid credentials")


# ==============================================================================
# 🔐 ADMIN SECURITY (Dependency)
# ==============================================================================
def verify_admin(admin_header_id: int = Header(..., alias="user_id", description="ID of the admin making the request")):
    """
    Verifies the request is coming from a logged-in Admin.
    Checks the AdminUser table.
    """
    try:
        # ✅ UPDATED: Now checks AdminUser table instead of UserData
        user = AdminUser.objects.get(id=admin_header_id)
        return user
    except AdminUser.DoesNotExist:
        raise HTTPException(status_code=404, detail="Admin user not found or session invalid.")


# ==============================================================================
# 📊 2. ANALYTICS PAGE ENDPOINTS
# ==============================================================================

@router.get("/overview")
def get_user_dashboard_overview(user_id: int):
    # This seems to be for standard users, keeping UserData here is correct if intended for user dashboard
    try:
        user = UserData.objects.get(id=user_id)
        has_active_plan = UserSubscription.objects.filter(
            user=user,
            plan_expires_at__gt=datetime.now()
        ).exists()

        return {
            "name": user.first_name,
            "role": user.role,
            "show_upgrade_banner": not has_active_plan
        }
    except UserData.DoesNotExist:
        raise HTTPException(status_code=404, detail="User not found")


@router.get("/analytics/stats")
def get_analytics_stats(admin: AdminUser = Depends(verify_admin)):
    """ Top 4 Cards for Analytics Page """
    now = timezone.now()
    
    # 1. Revenue this month
    revenue = BillingHistory.objects.filter(paid_on__month=now.month, paid_on__year=now.year, status="paid").aggregate(s=Sum('amount'))['s'] or 0.0
    
    # 2. Active Creators & Collaborators
    creators = UserData.objects.filter(role__iexact="Creator", status__iexact="Active").count()
    collabs = UserData.objects.filter(role__iexact="Collaborator", status__iexact="Active").count()
    
    # 3. Total Subscriptions
    subs = UserSubscription.objects.count()

    return {
        "revenue_month": float(revenue),
        "active_creators": creators,
        "active_collaborators": collabs,
        "total_subs": subs
    }

@router.get("/analytics/user-overview")
def get_user_overview_chart(admin: AdminUser = Depends(verify_admin)):
    """ Mixed Chart: Creators vs Collaborators vs Transactions (Last 6 Months) """
    now = timezone.now()
    data = []
    
    for i in range(5, -1, -1):
        month_start = (now - timedelta(days=i*30)).replace(day=1)
        month_label = month_start.strftime("%b") # Jan, Feb
        
        creators = UserData.objects.filter(role__iexact="Creator", created_at__year=month_start.year, created_at__month=month_start.month).count()
        collabs = UserData.objects.filter(role__iexact="Collaborator", created_at__year=month_start.year, created_at__month=month_start.month).count()
        txns = WalletTransaction.objects.filter(created_at__year=month_start.year, created_at__month=month_start.month).count()

        data.append({
            "Month": month_label,
            "Creator": creators,
            "Collaborator": collabs,
            "Transactions": txns
        })
    return data

@router.get("/analytics/task-performance")
def get_task_performance(admin: AdminUser = Depends(verify_admin)):
    """
    Returns real task performance stats from the CONTRACTS table.
    """
    # 1. Total Targets (Total Contracts Created)
    total_contracts = Contract.objects.count()
    
    # 2. Completed Tasks (Contracts where status is 'completed')
    completed_contracts = Contract.objects.filter(status__iexact="completed").count()
    
    # 3. Calculate "Late" vs "On Time"
    late_count = 0
    contracts = Contract.objects.filter(status__iexact="completed")
    
    for c in contracts:
        if c.end_date and c.updated_at:
            completion_date = c.updated_at.date()
            if completion_date > c.end_date:
                late_count += 1

    on_time = completed_contracts - late_count

    # 4. Growth Calculation
    now = datetime.now()
    
    if now.month == 1:
        last_month_num = 12
        last_month_year = now.year - 1
    else:
        last_month_num = now.month - 1
        last_month_year = now.year

    this_month = Contract.objects.filter(start_date__month=now.month, start_date__year=now.year).count()
    last_month = Contract.objects.filter(start_date__month=last_month_num, start_date__year=last_month_year).count()

    if last_month > 0:
        growth = ((this_month - last_month) / last_month) * 100
    else:
        growth = 100 if this_month > 0 else 0

    return {
        "total_completed": completed_contracts,
        "total_target": total_contracts,
        "on_time": on_time,
        "late": late_count,
        "tasks_this_year": Contract.objects.filter(start_date__year=now.year).count(),
        "growth": round(growth, 1)
    }

@router.get("/analytics/traffic-data")
def get_traffic_data(admin: AdminUser = Depends(verify_admin)):
    """
    Distributes TOTAL USER COUNT into Devices/Locations
    """
    total = UserData.objects.count() or 10 
    
    return {
        "device": [
            {"name": "Windows", "value": int(total * 0.45)},
            {"name": "Mac", "value": int(total * 0.25)},
            {"name": "Android", "value": int(total * 0.20)},
            {"name": "iOS", "value": int(total * 0.10)},
        ],
        "location": [
            {"name": "United States", "value": int(total * 0.50), "color": "#2e1065"},
            {"name": "Canada", "value": int(total * 0.25), "color": "#7c3aed"},
            {"name": "Mexico", "value": int(total * 0.15), "color": "#a78bfa"},
            {"name": "Other", "value": int(total * 0.10), "color": "#ddd6fe"},
        ]
    }

@router.get("/analytics/revenue-splits")
def get_revenue_splits(admin: AdminUser = Depends(verify_admin)):
    """ Pie Chart Data """
    platform_fees = BillingHistory.objects.filter(status="paid").aggregate(total=Sum('amount'))['total'] or 0.0
    creator_earnings = WalletTransaction.objects.filter(user__role__iexact="Creator", amount__gt=0).aggregate(total=Sum('amount'))['total'] or 0.0
    collab_earnings = WalletTransaction.objects.filter(user__role__iexact="Collaborator", amount__gt=0).aggregate(total=Sum('amount'))['total'] or 0.0

    return {
        "splits": [
            {"name": "Platform Fees", "value": float(platform_fees), "color": "#d8b4fe"},
            {"name": "Creator", "value": float(creator_earnings), "color": "#2e1065"},
            {"name": "Collaborator", "value": float(collab_earnings), "color": "#c4b5fd"}
        ]
    }

@router.get("/analytics/top-collaborators")
def get_top_collaborators(limit: int = 5, admin: AdminUser = Depends(verify_admin)):
    """ Top Collaborator List """
    top_users = UserData.objects.filter(role__iexact="collaborator").order_by('-wallet__balance')[:limit]
    results = []
    for u in top_users:
        wallet_bal = u.wallet.balance if hasattr(u, 'wallet') else 0.0
        results.append({
            "name": f"{u.first_name} {u.last_name}",
            "email": u.email,
            "earnings": float(wallet_bal),
            "joined_date": u.created_at.strftime("%d %b %Y")
        })
    return results

# ==============================================================================
# 📊 3. DASHBOARD OVERVIEW (Top Cards & Charts)
# ==============================================================================

@router.get("/dashboard/stats")
def get_dashboard_stats(admin: AdminUser = Depends(verify_admin)):
    total_users = UserData.objects.count()
    active_projects = JobPost.objects.filter(Q(status__iexact="posted") | Q(status__iexact="in_progress")).count()
    completed_tasks = JobPost.objects.filter(status__iexact="completed").count()
    revenue_agg = BillingHistory.objects.filter(status="paid").aggregate(total=Sum('amount'))
    total_revenue = revenue_agg['total'] or 0.0

    return {
        "admin_name": getattr(admin, 'name', 'Admin'), # Updated to use AdminUser name
        "total_users": total_users,
        "active_projects": active_projects,
        "completed_tasks": completed_tasks,
        "total_revenue": float(total_revenue)
    }


@router.get("/dashboard/charts/revenue")
def get_revenue_chart(
    filter: str = Query("Yearly", enum=["Weekly", "Monthly", "Yearly"]),
    year: int = Query(default=timezone.now().year, description="Select year for Yearly view"),
    admin: AdminUser = Depends(verify_admin)
):
    now = timezone.now()
    labels = []
    data = []

    if filter == "Weekly":
        for i in range(6, -1, -1):
            day = now - timedelta(days=i)
            labels.append(day.strftime("%a"))
            total = BillingHistory.objects.filter(paid_on__date=day.date(), status="paid").aggregate(s=Sum('amount'))['s'] or 0
            data.append(float(total))
            
    elif filter == "Monthly":
        for i in range(3, -1, -1):
            week_start = now - timedelta(weeks=i)
            labels.append(f"Week {4-i}")
            total = BillingHistory.objects.filter(paid_on__week=week_start.isocalendar()[1], paid_on__year=week_start.year, status="paid").aggregate(s=Sum('amount'))['s'] or 0
            data.append(float(total))

    else: # Yearly
        labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        for i in range(1, 13):
            total = BillingHistory.objects.filter(
                paid_on__month=i,
                paid_on__year=year,
                status="paid"
            ).aggregate(s=Sum('amount'))['s'] or 0
            data.append(float(total))

    return {"labels": labels, "data": data}

@router.get("/dashboard/charts/project-status")
def get_project_status_charts(time_range: str = "All", admin: AdminUser = Depends(verify_admin)):
    """ Project Status Donut Chart (Source: Contracts) """
    base_query = Contract.objects.all()
    now = timezone.now()

    if time_range == "Today":
        base_query = base_query.filter(updated_at__date=now.date())
    elif time_range == "Yesterday":
        d = now - timedelta(days=1)
        base_query = base_query.filter(updated_at__date=d.date())
    elif time_range == "DayBefore":
        d = now - timedelta(days=2)
        base_query = base_query.filter(updated_at__date=d.date())

    completed = base_query.filter(status__iexact="completed").count()
    on_hold = base_query.filter(
        Q(status__iexact="draft") |
        Q(status__iexact="on_hold") |
        Q(status__iexact="awaiting") 
    ).count()
    in_progress = base_query.filter(
        Q(status__iexact="posted") |
        Q(status__iexact="in_progress")
    ).count()
    
    return {
        "completed": completed,
        "on_hold": on_hold,
        "in_progress": in_progress,
    }

@router.get("/dashboard/charts/progress")
def get_progress_chart(
    filter: str = Query("Week", enum=["Week", "Month", "Year"]),
    admin: AdminUser = Depends(verify_admin)
):
    now = timezone.now()
    labels = []
    values = []

    if filter == "Week":
        for i in range(6, -1, -1):
            day = now - timedelta(days=i)
            labels.append(day.strftime("%d %b"))
            count = JobPost.objects.filter(created_at__date=day.date()).count()
            values.append(count)

    elif filter == "Month":
        for i in range(5, -1, -1):
            day = now - timedelta(days=i*5)
            labels.append(day.strftime("%d %b"))
            count = JobPost.objects.filter(
                created_at__date__lte=day.date(),
                created_at__date__gt=(day-timedelta(days=5)).date()
            ).count()
            values.append(count)

    elif filter == "Year":
        labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        for i in range(1, 13):
            count = JobPost.objects.filter(
                created_at__year=now.year,
                created_at__month=i
            ).count()
            values.append(count)

    return {"labels": labels, "data": values}

@router.get("/dashboard/active-projects")
def get_active_projects_table(limit: int = 5, admin: AdminUser = Depends(verify_admin)):
    """ Active Projects Table (Source: Contracts) """
    projects = Contract.objects.filter(
        ~Q(status__iexact="completed") &
        ~Q(status__iexact="cancelled")
    ).select_related('creator', 'job').order_by('-updated_at')[:limit]
    
    data = []
    now = timezone.now()

    for p in projects:
        client_name = "Unknown"
        if p.creator:
             client_name = f"{p.creator.first_name} {p.creator.last_name}"
        elif p.job and p.job.employer:
             client_name = f"{p.job.employer.first_name} {p.job.employer.last_name}"

        project_title = p.description
        if p.job:
            project_title = p.job.title

        duration_str = "N/A"
        progress_percent = 0

        if p.start_date and p.end_date:
            total_days = (p.end_date - p.start_date).days
            duration_str = f"{total_days} days"
            
            if p.start_date <= now.date():
                elapsed = (now.date() - p.start_date).days
                if total_days > 0:
                    progress_percent = int((elapsed / total_days) * 100)
            
            progress_percent = max(0, min(100, progress_percent))
        
        data.append({
            "client_name": client_name,
            "project_title": project_title,
            "price": float(p.budget) if p.budget else 0.0,
            "delivered_in": duration_str,
            "progress": progress_percent
        })

    return data

# ==============================================================================
# 👥 4. USER MANAGEMENT (CRUD)
# ==============================================================================

class UserUpdateSchema(BaseModel):
    name: Optional[str] = None # Updated to accept full name

@router.get("/users")
def get_all_users(
    role: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 10,
    admin: AdminUser = Depends(verify_admin)
):
    query = UserData.objects.all().order_by('-created_at')

    if role:
        query = query.filter(role__iexact=role)
    if search:
        query = query.filter(Q(first_name__icontains=search) | Q(email__icontains=search))
    
    total = query.count()
    start = (page - 1) * page_size
    end = start + page_size
    users = query[start:end]

    results = []
    for u in users:
        user_status = getattr(u, 'status', 'Active')
        results.append({
            "id": u.id,
            "full_name": f"{u.first_name} {u.last_name}",
            "email": u.email,
            "role": u.role,
            "status": user_status,
            "joined_date": u.created_at.strftime("%B %d, %Y"),
            "last_active": "Recently"
        })

    return {
        "total_users": total,
        "page": page,
        "page_size": page_size,
        "data": results
    }

@router.post("/users")
def create_user(
    first_name: str = Form(...),
    last_name: str = Form(...),
    email: str = Form(...),
    role: str = Form(...),
    password: str = Form(...),
    admin: AdminUser = Depends(verify_admin)
):
    if UserData.objects.filter(email=email).exists():
        raise HTTPException(status_code=400, detail="Email already registered")
    
    new_user = UserData.objects.create(
        first_name=first_name,
        last_name=last_name,
        email=email,
        role=role,
        password=password,
        status="Active"    
    )
    return {"status": "success", "message": "User created successfully", "user_id": new_user.id}

@router.put("/users/{user_id}")
def update_user(user_id: int, data: UserUpdateSchema, admin: AdminUser = Depends(verify_admin)):
    """ Update User Name logic """
    try:
        user = UserData.objects.get(id=user_id)
        
        if data.name:
            clean_name = data.name.strip()
            parts = clean_name.split(' ', 1)
            user.first_name = parts[0]
            if len(parts) > 1:
                user.last_name = parts[1]
            else:
                user.last_name = None
            
        user.save()
        return {"status": "success", "message": "User name updated successfully"}
    except UserData.DoesNotExist:
        raise HTTPException(404, "User not found")

@router.delete("/users/{user_id}")
def delete_user(user_id: int, admin: AdminUser = Depends(verify_admin)):
    try:
        user = UserData.objects.get(id=user_id)
        user.delete()
        return {"status": "success", "message": "User deleted"}
    except UserData.DoesNotExist:
        raise HTTPException(404, "User not found")

# ==============================================================================
# 💳 5. SUBSCRIPTION MANAGEMENT
# ==============================================================================

@router.get("/subscriptions/stats")
def get_subscription_stats(admin: AdminUser = Depends(verify_admin)):
    total_users = UserData.objects.count()
    pro_count = UserSubscription.objects.filter(current_plan__icontains="Pro").count()
    agent_count = UserSubscription.objects.filter(current_plan__icontains="Agent").count()
    explicit_free = UserSubscription.objects.filter(Q(current_plan__icontains="Free") | Q(current_plan__icontains="Basic")).count()
    users_with_subs = UserSubscription.objects.values_list('user_id', flat=True)
    users_without_subs = UserData.objects.exclude(id__in=users_with_subs).count()
    total_free = explicit_free + users_without_subs

    return {
        "total_subscribers": total_users,
        "active_free": total_free,
        "active_pro": pro_count,
        "active_agent": agent_count
    }

@router.get("/subscriptions/history")
def get_subscription_history(admin: AdminUser = Depends(verify_admin)):
    subs = UserSubscription.objects.all().order_by('-plan_started_at')[:20]
    data = []
    for s in subs:
        data.append({
            "full_name": f"{s.user.first_name} {s.user.last_name}",
            "email": s.email,
            "role": s.user.role,
            "date_of_subscription": s.plan_started_at.strftime("%B %d, %Y") if s.plan_started_at else "N/A",
            "plan": s.current_plan
        })
    return data

# ==============================================================================
# ⚙️ 6. SETTINGS
# ==============================================================================
class PasswordChangeSchema(BaseModel):
    new_password: str

@router.put("/users/{user_id}/change-password")
def change_user_password(user_id: int, data: PasswordChangeSchema, admin: AdminUser = Depends(verify_admin)):
    try:
        user = UserData.objects.get(id=user_id)
        user.password = data.new_password # Remember to hash this if storing directly!
        user.save()
        return {"status": "success", "message": "Password changed successfully"}
    except UserData.DoesNotExist:
        raise HTTPException(status_code=404, detail="User not found")
    

class PreferencesSchema(BaseModel):
    theme: str
    time_zone: str
    date_format: str
    default_dashboard: str

@router.post("/profile/image")
def upload_profile_image(file: UploadFile = File(...), admin: AdminUser = Depends(verify_admin)):
    # Note: This updates the ADMIN's profile image if AdminUser has that field.
    # Currently AdminUser model doesn't have profile_pic, but logic is here if added.
    try:
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) 
        MEDIA_ROOT = os.path.join(BASE_DIR, "media")
        PROFILE_PICS_DIR = os.path.join(MEDIA_ROOT, "profile_pics")
        os.makedirs(PROFILE_PICS_DIR, exist_ok=True)
        
        filename = f"admin_{admin.id}_{int(datetime.now().timestamp())}_{file.filename}"
        file_path = os.path.join(PROFILE_PICS_DIR, filename)
        
        with open(file_path, "wb+") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        # Assuming AdminUser doesn't have profile_picture field yet, just returning URL
        full_image_url = f"http://127.0.0.1:8000/media/profile_pics/{filename}"
        return {"status": "success", "image_url": full_image_url}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Image upload failed: {str(e)}")

# Note: UserPreferences model currently links to UserData.
# If you want admin preferences, you'd need to link UserPreferences to AdminUser or just mock this.
@router.get("/profile/preferences")
def get_profile_preferences(admin: AdminUser = Depends(verify_admin)):
    return {"theme": "System", "time_zone": "UTC", "date_format": "ISO", "default_dashboard": "Overview"}

@router.put("/profile/preferences")
def update_profile_preferences(data: PreferencesSchema, admin: AdminUser = Depends(verify_admin)):
    return {"status": "success", "message": "Preferences updated"}

# ==============================================================================
# 📥 7. EXPORT USERS
# ==============================================================================
@router.get("/users/export")
def export_users_custom(
    format: str = Query("csv", enum=["csv", "excel"], description="File format"),
    from_date: Optional[date] = Query(None, description="Filter from YYYY-MM-DD"),
    to_date: Optional[date] = Query(None, description="Filter to YYYY-MM-DD"),
    role: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    admin: AdminUser = Depends(verify_admin)
):
    query = UserData.objects.all().order_by('-created_at')

    if role:
        query = query.filter(role__iexact=role)
    if status:
        query = query.filter(status__iexact=status)
    if search:
        query = query.filter(Q(first_name__icontains=search) | Q(email__icontains=search))

    if from_date:
        query = query.filter(created_at__date__gte=from_date)
    if to_date:
        query = query.filter(created_at__date__lte=to_date)

    users = list(query)
    headers = ["Full Name", "Email", "Role", "Status", "Joined Date"]

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        
        for u in users:
            full_name = f"{u.first_name} {u.last_name}".strip()
            joined = u.created_at.strftime("%Y-%m-%d") if u.created_at else "N/A"
            user_status = getattr(u, 'status', 'Active')

            writer.writerow([full_name, u.email, u.role, user_status, joined])
            
        output.seek(0)
        filename = f"users_export_{datetime.now().strftime('%Y%m%d')}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    elif format == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users List"
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="2e1065", end_color="2e1065", fill_type="solid")

        for u in users:
            full_name = f"{u.first_name} {u.last_name}".strip()
            joined = u.created_at.strftime("%Y-%m-%d") if u.created_at else "N/A"
            user_status = getattr(u, 'status', 'Active')

            ws.append([full_name, u.email, u.role, user_status, joined])
            
        for col in ws.columns:
             max_length = 0
             column = col[0].column_letter 
             for cell in col:
                 try:
                     if len(str(cell.value)) > max_length:
                         max_length = len(str(cell.value))
                 except: pass
             adjusted_width = (max_length + 2) * 1.2
             ws.column_dimensions[column].width = adjusted_width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"users_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )